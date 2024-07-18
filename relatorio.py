import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt

# Função para processar o arquivo carregado
def processar_arquivo(caminho_arquivo, caminho_salvar):
    try:
        # Ler o arquivo carregado
        planilha = pd.read_csv(caminho_arquivo)

        # Renomear as colunas para o português
        planilha = planilha.rename(columns={'Date': 'Data', 
                                            'Source': 'Origem', 
                                            'Ring Group': 'Grupo de Chamada', 
                                            'Destination': 'Destino', 
                                            'Duration': 'Duração'})

        # Remover linhas onde o destino é '400', 's', 't', '1'
        planilha = planilha[~planilha['Destino'].isin(['400', 's', 't', '1'])]

        # Verificar se os valores que você deseja substituir estão presentes na coluna "Status"
        valores_substituir = {'ANSWERED': 'RESPONDIDAS', 'NO ANSWER': 'NÃO RESPONDIDAS', 'BUSY': 'OCUPADO', 'FAILED': 'FALHOU', 'CONGESTION': 'CONGESTIONADO'}
        planilha['Status'] = planilha['Status'].replace(valores_substituir)

        # Verificar as condições para determinar o valor da coluna "Tipo" baseado na coluna "Origem" 
        # Para informar se a ligação é Efetuada ou Recebida
        # A ligação só é efetuada por um ramal, logo a origem tem que ser um número de 4 dígitos
        def determinar_tipo(row):
            if len(str(row['Origem'])) == 4 or row['Origem'] == 7130327062:
                return 'Efetuada'
            else:
                return 'Recebida'

        # Aplicar a função para determinar o valor da coluna "Tipo"
        planilha['Tipo'] = planilha.apply(determinar_tipo, axis=1)

        # Remover colunas desnecessárias
        colunas_para_remover = ['Src. Channel', 'Account Code', 'Dst. Channel', 'UniqueID', 
                                'Recording', 'Cnum', 'Cnam', 'Outbound Cnum', 'DID', 'User Field']
        planilha_sem_colunas = planilha.drop(columns=colunas_para_remover)

        # Redefinir o índice das colunas
        planilha_sem_colunas = planilha_sem_colunas.loc[:, ~planilha_sem_colunas.columns.str.contains('^Unnamed')]

        # Salvar o arquivo processado
        planilha_sem_colunas.to_excel(caminho_salvar, index=False)

        # Adicionar um gráfico na planilha
        adicionar_grafico(caminho_salvar, planilha_sem_colunas)

        # Mostrar mensagem de sucesso
        messagebox.showinfo("Sucesso", f"Planilha salva com sucesso em: {caminho_salvar}")
    except Exception as error:
        messagebox.showerror("Erro", f"Ocorreu um erro durante o processo: {str(error)}")

# Função para adicionar um gráfico na planilha
def adicionar_grafico(caminho_salvar, planilha):
    # Criar um gráfico de exemplo
    grafico = planilha['Status'].value_counts().plot(kind='bar', title='Status das Chamadas')
    plt.xticks(rotation=0)  # Definir a rotação dos rótulos do eixo X para 0 (horizontal)
    plt.tight_layout()
    
    # Salvar o gráfico como uma imagem
    caminho_imagem = "grafico.png"
    plt.savefig(caminho_imagem)
    plt.close()

    # Carregar a planilha existente
    workbook = load_workbook(caminho_salvar)
    # Adicionar uma nova aba
    aba_grafico = workbook.create_sheet(title="Gráfico")

    # Inserir a imagem do gráfico na nova aba
    img = Image(caminho_imagem)
    aba_grafico.add_image(img, 'A1')

    # Salvar a planilha com a nova aba
    workbook.save(caminho_salvar)

# Função para abrir o diálogo de seleção de arquivo
def local_arquivo():
    caminho_arquivo = filedialog.askopenfilename()
    arquivo_var.set(caminho_arquivo)
    return caminho_arquivo

# Função para abrir o diálogo de seleção de pasta para salvar o arquivo
def local_salvar():
    caminho_salvar = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    salvar_var.set(caminho_salvar)
    return caminho_salvar

# Criar a janela principal
janela = tk.Tk()
janela.title("Análise de planilha telefônica")
janela.geometry("600x250")

# Impedir que a janela seja redimensionada
janela.resizable(False, False)

# Estilo ttk
style = ttk.Style()
style.configure("TLabel", padding=6, font=('Helvetica', 12))
style.configure("TButton", padding=6, font=('Helvetica', 12))
style.configure("TEntry", padding=6, font=('Helvetica', 12))

# Descrição e campo de entrada para o caminho do arquivo
label_arquivo = ttk.Label(janela, text="Informe a planilha exportada do sistema:")
label_arquivo.grid(row=0, column=0, padx=10, pady=5, sticky="W")

frame_arquivo = ttk.Frame(janela)
frame_arquivo.grid(row=1, column=0, padx=10, pady=5, sticky="W")

arquivo_var = tk.StringVar()
entrada_arquivo = ttk.Entry(frame_arquivo, textvariable=arquivo_var, width=50)
entrada_arquivo.pack(side=tk.LEFT, padx=5)

botao_procurar_arquivo = ttk.Button(frame_arquivo, text="Procurar Arquivo", command=local_arquivo)
botao_procurar_arquivo.pack(side=tk.LEFT)

# Descrição e campo de entrada para o caminho de salvar o arquivo
label_salvar = ttk.Label(janela, text="Escolha onde salvar a planilha modificada:")
label_salvar.grid(row=2, column=0, padx=10, pady=5, sticky="W")

frame_salvar = ttk.Frame(janela)
frame_salvar.grid(row=3, column=0, padx=10, pady=5, sticky="W")

salvar_var = tk.StringVar()
entrada_salvar = ttk.Entry(frame_salvar, textvariable=salvar_var, width=50)
entrada_salvar.pack(side=tk.LEFT, padx=5)

botao_procurar_salvar = ttk.Button(frame_salvar, text="Salvar Como", command=local_salvar)
botao_procurar_salvar.pack(side=tk.LEFT)

# Botão para processar o arquivo
botao_processar = ttk.Button(janela, text="Processar", command=lambda: processar_arquivo(arquivo_var.get(), salvar_var.get()))
botao_processar.grid(row=4, column=0, padx=10, pady=20, sticky="W")

# Executar o loop de eventos do tkinter
janela.mainloop()